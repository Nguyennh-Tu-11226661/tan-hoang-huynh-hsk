import csv
from io import BytesIO

from django.contrib import admin
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from .models import ConsultationRequest, TrialLessonBooking


FORMULA_PREFIXES = ("=", "+", "-", "@", "\t", "\r")


def spreadsheet_safe(value):
    if isinstance(value, str) and value.startswith(FORMULA_PREFIXES):
        return "'" + value
    return value


def consultation_rows(queryset):
    for item in queryset.select_related("course"):
        row = [
            item.full_name,
            item.phone,
            item.email,
            item.course.title if item.course else "",
            item.current_level,
            item.preferred_time,
            item.get_status_display(),
            timezone.localtime(item.created_at).strftime("%d/%m/%Y %H:%M"),
            item.message,
        ]
        yield [spreadsheet_safe(value) for value in row]


HEADERS = [
    "Họ và tên",
    "Số điện thoại",
    "Email",
    "Khóa học",
    "Trình độ",
    "Giờ liên hệ",
    "Trạng thái",
    "Ngày đăng ký",
    "Lời nhắn",
]


@admin.action(description="Xuất đăng ký đã chọn ra CSV")
def export_consultations_csv(modeladmin, request, queryset):
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = 'attachment; filename="dang-ky-tu-van.csv"'
    response["Cache-Control"] = "no-store, private"
    response.write("\ufeff")
    writer = csv.writer(response)
    writer.writerow(HEADERS)
    writer.writerows(consultation_rows(queryset))
    return response


@admin.action(description="Xuất đăng ký đã chọn ra Excel")
def export_consultations_excel(modeladmin, request, queryset):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Đăng ký tư vấn"
    sheet.append(HEADERS)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="C51D2D")
    for row in consultation_rows(queryset):
        sheet.append(row)
    widths = [24, 16, 28, 26, 18, 22, 16, 20, 40]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width
    stream = BytesIO()
    workbook.save(stream)
    response = HttpResponse(
        stream.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="dang-ky-tu-van.xlsx"'
    response["Cache-Control"] = "no-store, private"
    return response


@admin.register(ConsultationRequest)
class ConsultationRequestAdmin(admin.ModelAdmin):
    list_display = ("full_name", "phone", "course", "status", "created_at")
    list_filter = ("status", "course", "created_at")
    search_fields = ("full_name", "phone", "email")
    list_editable = ("status",)
    readonly_fields = (
        "created_at",
        "consent_given_at",
        "consent_version",
        "consent_text",
    )
    actions = [export_consultations_csv, export_consultations_excel]


@admin.register(TrialLessonBooking)
class TrialLessonBookingAdmin(admin.ModelAdmin):
    list_display = (
        "full_name",
        "phone",
        "booking_type",
        "preferred_date",
        "preferred_slot",
        "status",
    )
    list_filter = ("booking_type", "status", "preferred_date")
    search_fields = ("full_name", "phone", "email")
    list_editable = ("status",)
    readonly_fields = (
        "created_at",
        "consent_given_at",
        "consent_version",
        "consent_text",
    )
