import time
from datetime import timedelta
from io import BytesIO
from unittest.mock import patch

from django.core import signing
from django.core import mail
from django.core.management import call_command
from django.core.management.base import CommandError
from django.test import RequestFactory, TestCase
from django.test import override_settings
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from .admin import export_consultations_csv, export_consultations_excel
from .models import ConsultationRequest, TrialLessonBooking
from .notifications import notify_consultation_request


class AdmissionFormTests(TestCase):
    def token(self):
        return signing.dumps(int(time.time()) - 3)

    @override_settings(
        EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend",
        ADMISSION_NOTIFICATION_EMAILS=["tuvan@example.com"],
    )
    def test_consultation_submission(self):
        response = self.client.post(
            reverse("admissions:consultation"),
            {
                "full_name": "Nguyễn Minh An",
                "phone": "0988668596",
                "email": "an@example.com",
                "current_level": "Người mới",
                "preferred_time": "Buổi tối",
                "message": "Muốn học HSK",
                "consent": "on",
                "website": "",
                "form_started": self.token(),
            },
        )
        self.assertRedirects(response, reverse("admissions:success"))
        self.assertEqual(ConsultationRequest.objects.count(), 1)
        self.assertEqual(len(mail.outbox), 1)
        self.assertIn("Đăng ký tư vấn mới", mail.outbox[0].subject)

    def test_honeypot_blocks_spam(self):
        response = self.client.post(
            reverse("admissions:consultation"),
            {
                "full_name": "Spam Bot",
                "phone": "0988668596",
                "consent": "on",
                "website": "https://spam.example",
                "form_started": self.token(),
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(ConsultationRequest.objects.count(), 0)

    def test_trial_date_cannot_be_in_past(self):
        response = self.client.post(
            reverse("admissions:trial"),
            {
                "full_name": "Nguyễn Minh An",
                "phone": "0988668596",
                "booking_type": "trial",
                "preferred_date": (timezone.localdate() - timedelta(days=1)).isoformat(),
                "preferred_slot": "18:00–20:00",
                "consent": "on",
                "website": "",
                "form_started": self.token(),
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(TrialLessonBooking.objects.count(), 0)

    def test_trial_slot_must_be_valid_choice(self):
        response = self.client.post(
            reverse("admissions:trial"),
            {
                "full_name": "Nguyễn Minh An",
                "phone": "0988668596",
                "booking_type": "trial",
                "preferred_date": (timezone.localdate() + timedelta(days=1)).isoformat(),
                "preferred_slot": "=bad",
                "consent": "on",
                "website": "",
                "form_started": self.token(),
            },
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(TrialLessonBooking.objects.count(), 0)

    @override_settings(
        EMAIL_BACKEND="django.core.mail.backends.locmem.EmailBackend",
        ADMISSION_NOTIFICATION_EMAILS=["tuvan@example.com"],
    )
    def test_trial_submission_saves_booking(self):
        response = self.client.post(
            reverse("admissions:trial"),
            {
                "full_name": "Nguyễn Minh An",
                "phone": "0988668596",
                "email": "an@example.com",
                "booking_type": "placement",
                "preferred_date": (timezone.localdate() + timedelta(days=1)).isoformat(),
                "preferred_slot": "18:00–20:00",
                "current_level": "Mất gốc",
                "goal": "Muốn học HSK 3",
                "consent": "on",
                "website": "",
                "form_started": self.token(),
            },
        )

        self.assertRedirects(response, reverse("admissions:success"))
        self.assertEqual(TrialLessonBooking.objects.count(), 1)
        self.assertEqual(len(mail.outbox), 1)
        self.assertIn("Đặt lịch học thử/kiểm tra mới", mail.outbox[0].subject)

    def test_rate_limit_blocks_immediate_second_submission(self):
        payload = {
            "full_name": "Nguyễn Minh An",
            "phone": "0988668596",
            "consent": "on",
            "website": "",
            "form_started": self.token(),
        }
        self.client.post(reverse("admissions:consultation"), payload)
        response = self.client.post(
            reverse("admissions:consultation"),
            {**payload, "form_started": self.token()},
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Vui lòng chờ một phút")
        self.assertEqual(ConsultationRequest.objects.count(), 1)

    def test_sensitive_form_pages_are_not_cached(self):
        response = self.client.get(reverse("admissions:consultation"))
        self.assertEqual(response["Cache-Control"], "no-store, private")
        self.assertEqual(
            response["Permissions-Policy"],
            "camera=(), geolocation=(), microphone=()",
        )

    @override_settings(
        DEFAULT_FROM_EMAIL="website@example.com",
        ADMISSION_NOTIFICATION_EMAILS=["tuvan@example.com"],
    )
    @patch("admissions.notifications.send_mail", side_effect=RuntimeError("SMTP down"))
    def test_email_failure_is_logged(self, mocked_send_mail):
        request = ConsultationRequest.objects.create(
            full_name="Nguyễn Minh An",
            phone="0988668596",
            email="an@example.com",
        )

        with self.assertLogs("admissions.notifications", level="ERROR") as logs:
            result = notify_consultation_request(request)

        self.assertEqual(result, 0)
        self.assertTrue(mocked_send_mail.called)
        self.assertIn("Không gửi được email tuyển sinh", logs.output[0])


class ConsultationExportTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        ConsultationRequest.objects.create(
            full_name="Nguyễn Minh An",
            phone="0988668596",
            email="an@example.com",
            current_level="Người mới",
            preferred_time="Buổi tối",
            message="Muốn học HSK",
        )

    def setUp(self):
        self.request = RequestFactory().get("/quan-tri/")
        self.queryset = ConsultationRequest.objects.all()

    def test_csv_export(self):
        response = export_consultations_csv(None, self.request, self.queryset)
        text = response.content.decode("utf-8-sig")

        self.assertEqual(response.status_code, 200)
        self.assertIn("Nguyễn Minh An", text)
        self.assertIn("Số điện thoại", text)
        self.assertIn("dang-ky-tu-van.csv", response["Content-Disposition"])
        self.assertEqual(response["Cache-Control"], "no-store, private")

    def test_excel_export(self):
        response = export_consultations_excel(None, self.request, self.queryset)
        workbook = load_workbook(BytesIO(response.content), read_only=True)
        sheet = workbook["Đăng ký tư vấn"]

        self.assertEqual(response.status_code, 200)
        self.assertEqual(sheet["A2"].value, "Nguyễn Minh An")
        self.assertEqual(sheet["B2"].value, "0988668596")
        self.assertIn("dang-ky-tu-van.xlsx", response["Content-Disposition"])
        self.assertEqual(response["Cache-Control"], "no-store, private")

    def test_exports_escape_spreadsheet_formulas(self):
        ConsultationRequest.objects.create(
            full_name='=HYPERLINK("https://example.com")',
            phone="0988668597",
            email="formula@example.com",
            current_level="+cmd",
            preferred_time="@now",
            message="-danger",
        )

        response = export_consultations_csv(None, self.request, self.queryset)
        text = response.content.decode("utf-8-sig")
        self.assertIn("'=HYPERLINK", text)
        self.assertIn("'+cmd", text)

        response = export_consultations_excel(None, self.request, self.queryset)
        workbook = load_workbook(BytesIO(response.content), read_only=True)
        sheet = workbook["Đăng ký tư vấn"]
        exported_names = [sheet[f"A{row}"].value for row in range(2, 4)]
        self.assertIn("'=HYPERLINK(\"https://example.com\")", exported_names)


class PurgeAdmissionDataCommandTests(TestCase):
    def setUp(self):
        old_time = timezone.now() - timedelta(days=400)
        self.closed = ConsultationRequest.objects.create(
            full_name="Lead đã đóng",
            phone="0988668596",
            status=ConsultationRequest.Status.CLOSED,
        )
        self.active = ConsultationRequest.objects.create(
            full_name="Lead đang xử lý",
            phone="0988668597",
            status=ConsultationRequest.Status.CONTACTED,
        )
        self.booking = TrialLessonBooking.objects.create(
            full_name="Lịch đã hủy",
            phone="0988668598",
            booking_type=TrialLessonBooking.BookingType.TRIAL,
            preferred_date=timezone.localdate(),
            preferred_slot="18:00–20:00",
            status=TrialLessonBooking.Status.CANCELLED,
        )
        ConsultationRequest.objects.filter(pk__in=[self.closed.pk, self.active.pk]).update(
            created_at=old_time
        )
        TrialLessonBooking.objects.filter(pk=self.booking.pk).update(created_at=old_time)

    def test_dry_run_does_not_delete_records(self):
        call_command("purge_admission_data", "--dry-run", verbosity=0)
        self.assertEqual(ConsultationRequest.objects.count(), 2)
        self.assertEqual(TrialLessonBooking.objects.count(), 1)

    def test_confirm_deletes_only_old_terminal_records(self):
        call_command("purge_admission_data", "--confirm", verbosity=0)
        self.assertFalse(ConsultationRequest.objects.filter(pk=self.closed.pk).exists())
        self.assertTrue(ConsultationRequest.objects.filter(pk=self.active.pk).exists())
        self.assertFalse(TrialLessonBooking.objects.filter(pk=self.booking.pk).exists())

    def test_delete_requires_explicit_confirmation(self):
        with self.assertRaises(CommandError):
            call_command("purge_admission_data", verbosity=0)
